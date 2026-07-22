"""Gera modelo Excel para importação de ativos."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

EXEMPLOS = [
    ("Compressor AR-001", "AR-001"),
    ("Gerador GE-102", "GE-102"),
    ("Bomba Hidráulica BH-045", "BH-045"),
    ("Transformador TR-220", "TR-220"),
    ("Ar-condicionado Central AC-301", "AC-301"),
    ("Elevador EV-01", "EV-01"),
    ("Gerador de Emergência GE-200", "GE-200"),
    ("Porta Automática PA-12", "PA-12"),
    ("Câmara Fria CF-08", "CF-08"),
    ("Painel Elétrico PE-110", "PE-110"),
]

LINHAS_VAZIAS = 20

INSTRUCOES = [
    "Como usar este arquivo",
    "",
    "1. Preencha a aba Ativos com nome e codigo de cada equipamento.",
    "2. O codigo deve ser unico (nao repetir na planilha).",
    "3. Linhas em branco sao ignoradas na importacao.",
    "4. Se o codigo ja existir na filial, apenas o nome sera atualizado.",
    "5. Importe o arquivo .xlsx diretamente no sistema (Edicao > Importar).",
    "6. Tambem e possivel salvar como CSV UTF-8, se preferir.",
]


def write_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ativos"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")

    for col, title in enumerate(["nome", "codigo"], start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, (nome, codigo) in enumerate(EXEMPLOS, start=2):
        ws.cell(row=row_idx, column=1, value=nome)
        ws.cell(row=row_idx, column=2, value=codigo)

    start_blank = len(EXEMPLOS) + 2
    for row_idx in range(start_blank, start_blank + LINHAS_VAZIAS):
        ws.cell(row=row_idx, column=1, value="")
        ws.cell(row=row_idx, column=2, value="")

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.freeze_panes = "A2"

    inst = wb.create_sheet("Instrucoes")
    for row_idx, line in enumerate(INSTRUCOES, start=1):
        cell = inst.cell(row=row_idx, column=1, value=line)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
    inst.column_dimensions["A"].width = 72

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


if __name__ == "__main__":
    out = Path(__file__).resolve().parent.parent / "data" / "modelo_importacao_ativos.xlsx"
    write_xlsx(out)
    print(f"Arquivo criado: {out}")
