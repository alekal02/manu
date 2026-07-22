import csv
import io
import json
import os
import sqlite3
from datetime import datetime, timedelta

from werkzeug.security import check_password_hash, generate_password_hash

from db import get_conn, init_db

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LEGACY_ATIVOS = os.path.join(BASE_DIR, "data", "ativos.json")

LOCAIS = {
    "base": "Na base",
    "servico": "Em serviço externo",
    "deslocamento": "Em deslocamento",
    "terceiros": "Em terceiros",
}

LOCAIS_MANUTENCAO = {
    "base": "Manutenção na base",
    "terceiros": "Manutenção em terceiros",
}

LOCAIS_VALIDOS = tuple(LOCAIS.keys())
LOCAIS_MANUTENCAO_VALIDOS = tuple(LOCAIS_MANUTENCAO.keys())

TIPOS_EQUIPAMENTO = {
    "varredeira": "Varredeira",
    "caminhao_compactador": "Caminhão compactador",
    "rocadeira": "Roçadeira",
    "caminhao_pipa": "Caminhão pipa",
    "outros": "Outros",
}

TIPOS_VALIDOS = tuple(TIPOS_EQUIPAMENTO.keys())

FASES_OS = {
    "planejamento": "Planejamento",
    "diagnostico": "Diagnóstico",
    "execucao": "Execução",
    "divergencia_estoque": "Divergência de estoque",
    "aguardando_pecas": "Aguardando peças",
    "aguardando_terceiros": "Aguardando terceiros",
}

FASES_OS_VALIDAS = tuple(FASES_OS.keys())
FASE_OS_PADRAO = "planejamento"


def _row_to_dict(row):
    if row is None:
        return None
    d = dict(row)
    d["id"] = str(d["id"])
    if "base_id" in d and d["base_id"] is not None:
        d["base_id"] = str(d["base_id"])
    if "ativo_id" in d and d["ativo_id"] is not None:
        d["ativo_id"] = str(d["ativo_id"])
    for campo in ("em_manutencao", "ativa", "senha_alterada", "ativo", "aberta"):
        if campo in d and d[campo] is not None:
            d[campo] = bool(d[campo])
    return d


def normalizar_tipo(tipo):
    valor = (tipo or "").strip().lower().replace(" ", "_")
    return valor if valor in TIPOS_VALIDOS else "outros"


def _fmt_tipo(tipo):
    return TIPOS_EQUIPAMENTO.get(tipo, tipo or "—")


def _fmt_status(em_manutencao):
    return "Em manutenção" if em_manutencao else "Ativo"


def normalizar_fase(fase):
    valor = (fase or "").strip().lower().replace(" ", "_")
    return valor if valor in FASES_OS_VALIDAS else FASE_OS_PADRAO


def _fmt_fase(fase):
    return FASES_OS.get(fase, fase or "—")


def _texto_informado(valor):
    return bool((valor or "").strip())


def validar_fiscal_nao_remove_campo(valor_antigo, valor_novo, label):
    """Impede o fiscal de apagar conteúdo já registrado."""
    antigo = (valor_antigo or "").strip()
    novo = (valor_novo or "").strip()
    if _texto_informado(antigo) and not _texto_informado(novo):
        raise ValueError(f"O fiscal não pode remover {label} já informado.")
    if _texto_informado(antigo) and _texto_informado(novo) and len(novo) < len(antigo):
        raise ValueError(
            f"O fiscal só pode adicionar informações em {label}, não reduzir o texto."
        )


def mesclar_observacoes_fiscal(obs_antigo, obs_adicional):
    """Mantém observações anteriores e acrescenta novas entradas."""
    base = (obs_antigo or "").strip()
    extra = (obs_adicional or "").strip()
    if not extra:
        return base
    return f"{base}\n{extra}" if base else extra


def contar_bases():
    init_db()
    with get_conn() as conn:
        return conn.execute("SELECT COUNT(*) FROM bases").fetchone()[0]


def listar_bases(ativas_apenas=True):
    init_db()
    sql = "SELECT * FROM bases"
    if ativas_apenas:
        sql += " WHERE ativa = 1"
    sql += " ORDER BY codigo"
    with get_conn() as conn:
        rows = conn.execute(sql).fetchall()
    return [_row_to_dict(r) for r in rows]


def obter_base(base_id):
    init_db()
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM bases WHERE id = ?", (int(base_id),)).fetchone()
    return _row_to_dict(row)


def obter_base_resumo(base_id):
    init_db()
    with get_conn() as conn:
        usuarios = conn.execute(
            "SELECT COUNT(*) FROM usuarios WHERE base_id = ?",
            (int(base_id),),
        ).fetchone()[0]
        ativos = conn.execute(
            "SELECT COUNT(*) FROM ativos WHERE base_id = ?",
            (int(base_id),),
        ).fetchone()[0]
        historico = conn.execute(
            "SELECT COUNT(*) FROM historico_ativos WHERE base_id = ?",
            (int(base_id),),
        ).fetchone()[0]
    return {
        "total_usuarios": usuarios,
        "total_ativos": ativos,
        "total_historico": historico,
    }


def listar_bases_admin():
    bases = listar_bases(ativas_apenas=False)
    for base in bases:
        base.update(obter_base_resumo(base["id"]))
    return bases


def admin_atualizar_base(base_id, nome):
    nome = (nome or "").strip()
    if not nome:
        raise ValueError("O nome da filial é obrigatório.")

    base = obter_base(base_id)
    if not base:
        raise ValueError("Filial não encontrada.")

    with get_conn() as conn:
        conn.execute(
            "UPDATE bases SET nome = ? WHERE id = ?",
            (nome, int(base_id)),
        )


def admin_excluir_base(base_id):
    base = obter_base(base_id)
    if not base:
        raise ValueError("Filial não encontrada.")

    if contar_bases() <= 1:
        raise ValueError("Não é possível excluir a única filial do sistema.")

    bid = int(base_id)
    with get_conn() as conn:
        conn.execute("DELETE FROM manutencoes WHERE base_id = ?", (bid,))
        conn.execute("DELETE FROM historico_ativos WHERE base_id = ?", (bid,))
        conn.execute("DELETE FROM ativos WHERE base_id = ?", (bid,))
        conn.execute("DELETE FROM usuarios WHERE base_id = ?", (bid,))
        conn.execute("DELETE FROM bases WHERE id = ?", (bid,))

    return base


def autenticar(base_id, usuario, senha):
    init_db()
    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT * FROM usuarios
            WHERE base_id = ? AND usuario = ? AND ativo = 1
            """,
            (int(base_id), usuario.strip().lower()),
        ).fetchone()

    if row and check_password_hash(row["senha_hash"], senha):
        user = _row_to_dict(row)
        base = obter_base(base_id)
        user["base_nome"] = base["nome"] if base else ""
        user["base_codigo"] = base["codigo"] if base else ""
        return user
    return None


def obter_usuario(base_id, usuario):
    init_db()
    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT * FROM usuarios
            WHERE base_id = ? AND usuario = ? AND ativo = 1
            """,
            (int(base_id), usuario),
        ).fetchone()
    return _row_to_dict(row)


def atualizar_senha(base_id, usuario, nova_senha):
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE usuarios
            SET senha_hash = ?, senha_alterada = 1
            WHERE base_id = ? AND usuario = ?
            """,
            (generate_password_hash(nova_senha), int(base_id), usuario),
        )


def listar_ativos(base_id):
    init_db()
    with get_conn() as conn:
        rows = conn.execute(
            "SELECT * FROM ativos WHERE base_id = ? ORDER BY codigo",
            (int(base_id),),
        ).fetchall()
    return [_row_to_dict(r) for r in rows]


def obter_ativo(base_id, ativo_id):
    init_db()
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM ativos WHERE id = ? AND base_id = ?",
            (int(ativo_id), int(base_id)),
        ).fetchone()
    return _row_to_dict(row)


def _fmt_local(local):
    return LOCAIS.get(local, local or "Na base")


def normalizar_local(local):
    valor = (local or "base").strip().lower()
    return valor if valor in LOCAIS_VALIDOS else "base"


def registrar_historico(
    base_id,
    ativo_id,
    codigo,
    nome,
    acao,
    usuario,
    usuario_nome,
    detalhes,
    em_manutencao=None,
    local=None,
    ordem_servico=None,
    conn=None,
):
    agora = datetime.now().isoformat(timespec="seconds")
    params = (
        int(base_id),
        int(ativo_id) if ativo_id else None,
        codigo,
        nome,
        acao,
        usuario,
        usuario_nome or usuario,
        detalhes,
        int(em_manutencao) if em_manutencao is not None else None,
        local,
        ordem_servico or "",
        agora,
    )
    sql = """
            INSERT INTO historico_ativos
                (base_id, ativo_id, codigo, nome, acao, usuario, usuario_nome,
                 detalhes, em_manutencao, local, ordem_servico, criado_em)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
    if conn is not None:
        conn.execute(sql, params)
        return

    with get_conn() as db_conn:
        db_conn.execute(sql, params)


def _detalhes_edicao(antigo, novo):
    mudancas = []
    if bool(antigo["em_manutencao"]) != bool(novo["em_manutencao"]):
        mudancas.append(
            f"Status: {_fmt_status(antigo['em_manutencao'])} → {_fmt_status(novo['em_manutencao'])}"
        )
    if antigo.get("local") != novo.get("local"):
        mudancas.append(
            f"Local: {_fmt_local(antigo.get('local', 'base'))} → {_fmt_local(novo.get('local', 'base'))}"
        )
    if (antigo.get("ordem_servico") or "") != (novo.get("ordem_servico") or ""):
        mudancas.append(
            f"OS: '{antigo.get('ordem_servico') or '—'}' → '{novo.get('ordem_servico') or '—'}'"
        )
    if (antigo.get("observacoes") or "") != (novo.get("observacoes") or ""):
        mudancas.append("Observações atualizadas")
    if antigo.get("nome") != novo.get("nome"):
        mudancas.append(f"Nome: '{antigo['nome']}' → '{novo['nome']}'")
    if antigo.get("codigo") != novo.get("codigo"):
        mudancas.append(f"Código: '{antigo['codigo']}' → '{novo['codigo']}'")
    if (antigo.get("tipo") or "") != (novo.get("tipo") or ""):
        mudancas.append(
            f"Tipo: {_fmt_tipo(antigo.get('tipo'))} → {_fmt_tipo(novo.get('tipo'))}"
        )
    if (antigo.get("patrimonio") or "") != (novo.get("patrimonio") or ""):
        mudancas.append(
            f"Patrimônio: '{antigo.get('patrimonio') or '—'}' → '{novo.get('patrimonio') or '—'}'"
        )
    if (antigo.get("data_aquisicao") or "") != (novo.get("data_aquisicao") or ""):
        mudancas.append("Data de aquisição atualizada")
    return " · ".join(mudancas) if mudancas else "Registro atualizado"


def salvar_ativo(base_id, ativo_id, dados, usuario=None, usuario_nome=None, modo_fiscal=False):
    """Atualiza cadastro do equipamento (status via abrir/encerrar manutencao)."""
    antigo = obter_ativo(base_id, ativo_id)
    if not antigo:
        return

    nome = (dados.get("nome") or antigo["nome"]).strip()
    codigo = (dados.get("codigo") or antigo["codigo"]).strip()
    if not nome or not codigo:
        raise ValueError("Nome e código são obrigatórios.")

    tipo = normalizar_tipo(dados.get("tipo", antigo.get("tipo", "outros")))
    patrimonio = (
        dados.get("patrimonio") if "patrimonio" in dados else antigo.get("patrimonio") or ""
    )
    patrimonio = (patrimonio or "").strip()
    data_aquisicao = (
        dados.get("data_aquisicao")
        if "data_aquisicao" in dados
        else antigo.get("data_aquisicao") or ""
    )
    data_aquisicao = (data_aquisicao or "").strip() or None
    local = normalizar_local(dados.get("local", antigo.get("local", "base")))

    if modo_fiscal:
        validar_fiscal_nao_remove_campo(antigo.get("patrimonio"), patrimonio, "patrimônio")
        validar_fiscal_nao_remove_campo(
            antigo.get("data_aquisicao"), data_aquisicao, "data de aquisição"
        )
        observacoes = mesclar_observacoes_fiscal(
            antigo.get("observacoes"),
            dados.get("observacoes_adicionais", ""),
        )
    else:
        observacoes = (
            dados.get("observacoes")
            if "observacoes" in dados
            else antigo.get("observacoes") or ""
        )
        observacoes = (observacoes or "").strip()

    em_manutencao = bool(antigo["em_manutencao"])
    ordem_servico = antigo.get("ordem_servico") or ""

    novo = {
        "nome": nome,
        "codigo": codigo,
        "tipo": tipo,
        "patrimonio": patrimonio,
        "data_aquisicao": data_aquisicao or "",
        "em_manutencao": em_manutencao,
        "local": local,
        "ordem_servico": ordem_servico,
        "observacoes": observacoes,
    }
    agora = datetime.now().isoformat(timespec="seconds")

    try:
        with get_conn() as conn:
            conn.execute(
                """
                UPDATE ativos
                SET nome = ?, codigo = ?, tipo = ?, patrimonio = ?, data_aquisicao = ?,
                    local = ?, observacoes = ?, atualizado_em = ?
                WHERE id = ? AND base_id = ?
                """,
                (
                    novo["nome"],
                    novo["codigo"],
                    novo["tipo"],
                    novo["patrimonio"],
                    data_aquisicao,
                    novo["local"],
                    novo["observacoes"],
                    agora,
                    int(ativo_id),
                    int(base_id),
                ),
            )
            detalhes = _detalhes_edicao(antigo, novo)
            if not detalhes or detalhes == "Registro atualizado":
                detalhes = f"Cadastro atualizado · {_fmt_tipo(tipo)}"
            else:
                detalhes = f"Edição · {detalhes}"
            usr = (usuario or "").strip() or "fiscal"
            usr_nome = (usuario_nome or "").strip() or usr
            registrar_historico(
                base_id,
                ativo_id,
                novo["codigo"],
                novo["nome"],
                "edicao",
                usr,
                usr_nome,
                detalhes,
                em_manutencao=em_manutencao,
                local=local,
                ordem_servico=ordem_servico,
                conn=conn,
            )
    except sqlite3.IntegrityError as exc:
        raise ValueError("Já existe um equipamento com este código nesta filial.") from exc
    return novo


def excluir_ativo(base_id, ativo_id, usuario=None, usuario_nome=None):
    antigo = obter_ativo(base_id, ativo_id)
    if not antigo:
        return False

    with get_conn() as conn:
        conn.execute(
            "DELETE FROM manutencoes WHERE ativo_id = ? AND base_id = ?",
            (int(ativo_id), int(base_id)),
        )
        conn.execute(
            "DELETE FROM ativos WHERE id = ? AND base_id = ?",
            (int(ativo_id), int(base_id)),
        )

    if usuario:
        registrar_historico(
            base_id,
            None,
            antigo["codigo"],
            antigo["nome"],
            "exclusao",
            usuario,
            usuario_nome,
            f"Equipamento excluído · {antigo['codigo']} — {antigo['nome']}",
            em_manutencao=antigo["em_manutencao"],
            local=antigo.get("local"),
            ordem_servico=antigo.get("ordem_servico", ""),
        )

    return True


def criar_ativo(base_id, dados, usuario=None, usuario_nome=None):
    agora = datetime.now().isoformat(timespec="seconds")
    nome = (dados.get("nome") or "").strip()
    codigo = (dados.get("codigo") or "").strip()
    if not nome or not codigo:
        raise ValueError("Nome e código são obrigatórios.")

    tipo = normalizar_tipo(dados.get("tipo", "outros"))
    patrimonio = (dados.get("patrimonio") or "").strip()
    data_aquisicao = (dados.get("data_aquisicao") or "").strip() or None
    local = normalizar_local(dados.get("local", "base"))
    observacoes = (dados.get("observacoes") or "").strip()

    try:
        with get_conn() as conn:
            cur = conn.execute(
                """
                INSERT INTO ativos
                    (base_id, nome, codigo, tipo, patrimonio, data_aquisicao,
                     local, em_manutencao, ordem_servico, observacoes, atualizado_em)
                VALUES (?, ?, ?, ?, ?, ?, ?, 0, '', ?, ?)
                """,
                (
                    int(base_id),
                    nome,
                    codigo,
                    tipo,
                    patrimonio,
                    data_aquisicao,
                    local,
                    observacoes,
                    agora,
                ),
            )
            novo_id = str(cur.lastrowid)
            if usuario:
                registrar_historico(
                    base_id,
                    novo_id,
                    codigo,
                    nome,
                    "criacao",
                    usuario,
                    usuario_nome,
                    f"Equipamento cadastrado · {_fmt_tipo(tipo)} · Ativo",
                    em_manutencao=False,
                    local=local,
                    ordem_servico="",
                    conn=conn,
                )
    except sqlite3.IntegrityError as exc:
        raise ValueError("Já existe um equipamento com este código nesta filial.") from exc
    return novo_id


def _normalizar_campo_importacao(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def _validar_linhas_importacao(registros):
    linhas = []
    codigos_vistos = set()
    for i, nome, codigo in registros:
        if not nome and not codigo:
            continue
        if not nome or not codigo:
            raise ValueError(f"Linha {i}: nome e codigo são obrigatórios.")
        codigo_upper = codigo.upper()
        if codigo_upper in codigos_vistos:
            raise ValueError(f"Linha {i}: código duplicado no arquivo ({codigo}).")
        codigos_vistos.add(codigo_upper)
        linhas.append({"nome": nome, "codigo": codigo})
    return linhas


def _ler_linhas_csv(conteudo_bytes):
    texto = conteudo_bytes.decode("utf-8-sig")
    leitor = csv.DictReader(io.StringIO(texto))

    if not leitor.fieldnames:
        raise ValueError("Arquivo CSV vazio ou sem cabeçalho.")

    campos = {c.strip().lower() for c in leitor.fieldnames}
    if "codigo" not in campos or "nome" not in campos:
        raise ValueError("O arquivo deve conter as colunas: nome, codigo")

    registros = []
    for i, row in enumerate(leitor, start=2):
        nome = _normalizar_campo_importacao(row.get("nome") or row.get("Nome"))
        codigo = _normalizar_campo_importacao(
            row.get("codigo") or row.get("Codigo") or row.get("Código")
        )
        registros.append((i, nome, codigo))
    return registros


def _detectar_tipo_arquivo(conteudo_bytes, nome_arquivo=""):
    ext = os.path.splitext(nome_arquivo or "")[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return "xlsx"
    if conteudo_bytes[:2] == b"PK":
        return "xlsx"
    return "csv"


def _ler_linhas_xlsx(conteudo_bytes):
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(conteudo_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(
            "Não foi possível ler a planilha Excel. Verifique se o arquivo .xlsx está válido."
        ) from exc
    try:
        ws = wb["Ativos"] if "Ativos" in wb.sheetnames else wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if not header:
            raise ValueError("Planilha vazia ou sem cabeçalho.")

        header_map = {}
        for idx, col in enumerate(header):
            key = _normalizar_campo_importacao(col).lower()
            if key in ("nome", "codigo", "código"):
                header_map["codigo" if key == "código" else key] = idx

        if "nome" not in header_map or "codigo" not in header_map:
            raise ValueError("A planilha deve conter as colunas: nome, codigo")

        registros = []
        for i, row in enumerate(rows_iter, start=2):
            row = row or ()
            nome_idx = header_map["nome"]
            codigo_idx = header_map["codigo"]
            nome = _normalizar_campo_importacao(row[nome_idx] if nome_idx < len(row) else None)
            codigo = _normalizar_campo_importacao(
                row[codigo_idx] if codigo_idx < len(row) else None
            )
            registros.append((i, nome, codigo))
        return registros
    finally:
        wb.close()


def _ler_linhas_arquivo(conteudo_bytes, nome_arquivo=""):
    if _detectar_tipo_arquivo(conteudo_bytes, nome_arquivo) == "xlsx":
        return _ler_linhas_xlsx(conteudo_bytes)
    try:
        return _ler_linhas_csv(conteudo_bytes)
    except UnicodeDecodeError as exc:
        raise ValueError(
            "Arquivo inválido. Envie um CSV UTF-8 ou uma planilha Excel (.xlsx)."
        ) from exc


def importar_ativos_csv(
    base_id, conteudo_bytes, substituir=False, usuario=None, usuario_nome=None, nome_arquivo=""
):
    registros = _ler_linhas_arquivo(conteudo_bytes, nome_arquivo)
    linhas = _validar_linhas_importacao(registros)

    if not linhas:
        raise ValueError("Nenhum ativo válido encontrado no arquivo.")

    agora = datetime.now().isoformat(timespec="seconds")
    inseridos = 0
    atualizados = 0
    base_int = int(base_id)

    with get_conn() as conn:
        if substituir:
            removidos = conn.execute(
                "SELECT COUNT(*) FROM ativos WHERE base_id = ?", (base_int,)
            ).fetchone()[0]
            conn.execute("DELETE FROM ativos WHERE base_id = ?", (base_int,))
            if usuario and removidos:
                registrar_historico(
                    base_id,
                    None,
                    "—",
                    "Importação CSV",
                    "importacao",
                    usuario,
                    usuario_nome,
                    f"Lista substituída — {removidos} ativo(s) removido(s) antes da importação",
                    conn=conn,
                )

        for item in linhas:
            existente = conn.execute(
                "SELECT id, nome FROM ativos WHERE base_id = ? AND codigo = ?",
                (base_int, item["codigo"]),
            ).fetchone()

            if existente:
                conn.execute(
                    """
                    UPDATE ativos SET nome = ?, atualizado_em = ?
                    WHERE id = ?
                    """,
                    (item["nome"], agora, existente["id"]),
                )
                atualizados += 1
                if usuario:
                    registrar_historico(
                        base_id,
                        existente["id"],
                        item["codigo"],
                        item["nome"],
                        "importacao",
                        usuario,
                        usuario_nome,
                        f"Nome atualizado via CSV (era: {existente['nome']})",
                        conn=conn,
                    )
            else:
                cur = conn.execute(
                    """
                    INSERT INTO ativos
                        (base_id, nome, codigo, local, em_manutencao,
                         ordem_servico, observacoes, atualizado_em)
                    VALUES (?, ?, ?, 'base', 0, '', '', ?)
                    """,
                    (base_int, item["nome"], item["codigo"], agora),
                )
                inseridos += 1
                if usuario:
                    registrar_historico(
                        base_id,
                        cur.lastrowid,
                        item["codigo"],
                        item["nome"],
                        "importacao",
                        usuario,
                        usuario_nome,
                        "Ativo incluído via importação CSV",
                        em_manutencao=False,
                        local="base",
                        conn=conn,
                    )

    if usuario and not substituir:
        pass  # registros individuais já gravados acima

    return {"inseridos": inseridos, "atualizados": atualizados, "total": len(linhas)}


def seed_bases(qtd=20):
    init_db()
    if contar_bases() > 0:
        return {"criadas": 0, "mensagem": "Bases já existem no banco."}

    senha_padrao = generate_password_hash("1234")
    bases_criadas = 0
    usuarios_criados = 0

    with get_conn() as conn:
        for i in range(1, qtd + 1):
            codigo = f"{i:02d}"
            cur = conn.execute(
                "INSERT INTO bases (codigo, nome, ativa) VALUES (?, ?, 1)",
                (codigo, f"Filial {codigo}"),
            )
            base_id = cur.lastrowid
            bases_criadas += 1

            for usuario, nome, nivel in [
                ("gerente", f"Gerente Filial {codigo}", "gerente"),
                ("fiscal", f"Fiscal Filial {codigo}", "fiscal"),
            ]:
                conn.execute(
                    """
                    INSERT INTO usuarios
                        (base_id, usuario, senha_hash, nome, nivel, senha_alterada, ativo)
                    VALUES (?, ?, ?, ?, ?, 0, 1)
                    """,
                    (base_id, usuario, senha_padrao, nome, nivel),
                )
                usuarios_criados += 1

    return {
        "criadas": bases_criadas,
        "usuarios": usuarios_criados,
        "mensagem": f"{bases_criadas} filiais criadas com gerente e fiscal cada.",
    }


def migrar_json_legacy(base_codigo="01"):
    init_db()
    with get_conn() as conn:
        base = conn.execute(
            "SELECT id FROM bases WHERE codigo = ?", (base_codigo,)
        ).fetchone()

    if not base:
        raise ValueError(f"Filial {base_codigo} não encontrada. Execute o seed primeiro.")

    base_id = base["id"]
    agora = datetime.now().isoformat(timespec="seconds")

    if os.path.exists(LEGACY_ATIVOS):
        with open(LEGACY_ATIVOS, "r", encoding="utf-8") as f:
            ativos = json.load(f)

        with get_conn() as conn:
            for a in ativos:
                conn.execute(
                    """
                    INSERT INTO ativos
                        (base_id, nome, codigo, local, em_manutencao,
                         ordem_servico, observacoes, atualizado_em)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(base_id, codigo) DO UPDATE SET
                        nome = excluded.nome,
                        local = excluded.local,
                        em_manutencao = excluded.em_manutencao,
                        ordem_servico = excluded.ordem_servico,
                        observacoes = excluded.observacoes,
                        atualizado_em = excluded.atualizado_em
                    """,
                    (
                        base_id,
                        a["nome"],
                        a["codigo"],
                        a.get("local", "base"),
                        int(a.get("em_manutencao", False)),
                        a.get("ordem_servico", ""),
                        a.get("observacoes", ""),
                        a.get("atualizado_em", agora),
                    ),
                )

    return {"base": base_codigo, "mensagem": "Migração concluída."}


# --- Relatórios ---


def _historico_dict(row):
    d = dict(row)
    d["id"] = str(d["id"])
    if "base_id" in d and d["base_id"] is not None:
        d["base_id"] = str(d["base_id"])
    if d.get("ativo_id"):
        d["ativo_id"] = str(d["ativo_id"])
    if d.get("em_manutencao") is not None:
        d["em_manutencao"] = bool(d["em_manutencao"])
    return d


ACAO_LABEL = {
    "criacao": "Cadastro",
    "edicao": "Edição",
    "verificacao": "Check-in diário",
    "manutencao_abertura": "Abertura de OS",
    "manutencao_encerramento": "Encerramento de OS",
    "manutencao_fase": "Alteração de fase da OS",
    "manutencao_anotacao": "Anotação na OS",
    "manutencao_atualizacao": "Atualização da OS aberta",
    "importacao": "Importação CSV",
    "exclusao": "Exclusão",
}


def atualizar_status_diario(base_id, ativo_id, dados, usuario=None, usuario_nome=None):
    """Check-in diário do fiscal: operacional ou manutenção (com OS + local)."""
    antigo = obter_ativo(base_id, ativo_id)
    if not antigo:
        raise ValueError("Ativo não encontrado.")

    em_manutencao = bool(dados.get("em_manutencao"))
    observacoes = (dados.get("observacoes") or "").strip()

    if em_manutencao:
        ordem_servico = (dados.get("ordem_servico") or "").strip()
        if not ordem_servico:
            raise ValueError("Informe a ordem de serviço (OS) quando o ativo estiver em manutenção.")
        local = (dados.get("local") or "").strip().lower()
        if local not in LOCAIS_MANUTENCAO_VALIDOS:
            raise ValueError("Informe onde está a manutenção: na base ou em terceiros.")
    else:
        ordem_servico = ""
        local = "base"

    novo = {
        "nome": antigo["nome"],
        "codigo": antigo["codigo"],
        "em_manutencao": em_manutencao,
        "local": local,
        "ordem_servico": ordem_servico,
        "observacoes": observacoes,
    }
    agora = datetime.now().isoformat(timespec="seconds")

    with get_conn() as conn:
        conn.execute(
            """
            UPDATE ativos
            SET em_manutencao = ?, local = ?, ordem_servico = ?,
                observacoes = ?, atualizado_em = ?
            WHERE id = ? AND base_id = ?
            """,
            (
                int(novo["em_manutencao"]),
                novo["local"],
                novo["ordem_servico"],
                novo["observacoes"],
                agora,
                int(ativo_id),
                int(base_id),
            ),
        )

    if usuario:
        detalhes = _detalhes_edicao(antigo, novo)
        if not detalhes or detalhes == "Registro atualizado":
            if em_manutencao:
                detalhes = (
                    f"Check-in · Em manutenção · OS {ordem_servico} · {_fmt_local(local)}"
                )
            else:
                detalhes = "Check-in · Operacional"

        registrar_historico(
            base_id,
            ativo_id,
            novo["codigo"],
            novo["nome"],
            "verificacao",
            usuario,
            usuario_nome,
            detalhes,
            em_manutencao=novo["em_manutencao"],
            local=novo["local"],
            ordem_servico=novo["ordem_servico"],
        )

    return novo


def relatorio_diario(base_id, data_ref=None, busca=None):
    init_db()
    if not data_ref:
        data_ref = datetime.now().strftime("%Y-%m-%d")

    inicio = f"{data_ref} 00:00:00"
    fim_dt = datetime.strptime(data_ref, "%Y-%m-%d").replace(hour=0, minute=0, second=0)
    fim = (fim_dt + timedelta(days=1)).strftime("%Y-%m-%d %H:%M:%S")

    sql = """
        SELECT * FROM historico_ativos
        WHERE base_id = ? AND criado_em >= ? AND criado_em < ?
    """
    params = [int(base_id), inicio, fim]
    busca = (busca or "").strip()
    if busca:
        like = f"%{busca}%"
        sql += " AND (codigo LIKE ? OR nome LIKE ? OR ordem_servico LIKE ? OR detalhes LIKE ? OR usuario LIKE ? OR usuario_nome LIKE ?)"
        params.extend([like, like, like, like, like, like])
    sql += " ORDER BY criado_em DESC"

    with get_conn() as conn:
        rows = conn.execute(sql, params).fetchall()

    registros = [_historico_dict(r) for r in rows]
    por_acao = {}
    for r in registros:
        por_acao[r["acao"]] = por_acao.get(r["acao"], 0) + 1

    return {
        "data": data_ref,
        "registros": registros,
        "total": len(registros),
        "por_acao": por_acao,
        "busca": busca,
    }


def relatorio_mensal(base_id, ano, mes, busca=None):
    init_db()
    mes_str = f"{int(mes):02d}"
    ano_str = str(int(ano))

    busca = (busca or "").strip()
    filtro_busca = ""
    params_base = [int(base_id), ano_str, mes_str]
    if busca:
        like = f"%{busca}%"
        filtro_busca = " AND (codigo LIKE ? OR nome LIKE ? OR ordem_servico LIKE ? OR detalhes LIKE ? OR usuario LIKE ? OR usuario_nome LIKE ?)"
        params_base.extend([like, like, like, like, like, like])

    with get_conn() as conn:
        rows = conn.execute(
            f"""
            SELECT * FROM historico_ativos
            WHERE base_id = ?
              AND strftime('%Y', criado_em) = ?
              AND strftime('%m', criado_em) = ?
              {filtro_busca}
            ORDER BY criado_em DESC
            """,
            params_base,
        ).fetchall()

        por_dia_rows = conn.execute(
            f"""
            SELECT date(criado_em) AS dia, COUNT(*) AS total
            FROM historico_ativos
            WHERE base_id = ?
              AND strftime('%Y', criado_em) = ?
              AND strftime('%m', criado_em) = ?
              {filtro_busca}
            GROUP BY date(criado_em)
            ORDER BY dia DESC
            """,
            params_base,
        ).fetchall()

        counts = conn.execute(
            """
            SELECT COUNT(*), COALESCE(SUM(em_manutencao), 0)
            FROM ativos WHERE base_id = ?
            """,
            (int(base_id),),
        ).fetchone()

    registros = [_historico_dict(r) for r in rows]
    por_acao = {}
    ativos_alterados = set()
    for r in registros:
        por_acao[r["acao"]] = por_acao.get(r["acao"], 0) + 1
        if r.get("codigo") and r["codigo"] != "—":
            ativos_alterados.add(r["codigo"])

    total_ativos = counts[0]
    em_manutencao = counts[1] or 0

    por_dia = [{"dia": row["dia"], "total": row["total"]} for row in por_dia_rows]

    return {
        "ano": int(ano),
        "mes": int(mes),
        "registros": registros,
        "total": len(registros),
        "por_acao": por_acao,
        "por_dia": por_dia,
        "dias_com_atividade": len(por_dia),
        "ativos_movimentados": len(ativos_alterados),
        "busca": busca,
        "snapshot": {
            "total_ativos": total_ativos,
            "operacionais": total_ativos - em_manutencao,
            "manutencao": em_manutencao,
        },
    }


def obter_historico(historico_id):
    init_db()
    with get_conn() as conn:
        row = conn.execute(
            "SELECT * FROM historico_ativos WHERE id = ?",
            (int(historico_id),),
        ).fetchone()
    return _historico_dict(row) if row else None


def listar_historico_ativo(base_id, ativo_id, limite=80, conn=None):
    """Trilha de auditoria do fiscal para um equipamento."""
    sql = """
        SELECT * FROM historico_ativos
        WHERE base_id = ? AND ativo_id = ?
        ORDER BY criado_em DESC, id DESC
        LIMIT ?
    """
    params = (int(base_id), int(ativo_id), int(limite))
    if conn is not None:
        rows = conn.execute(sql, params).fetchall()
    else:
        init_db()
        with get_conn() as db:
            rows = db.execute(sql, params).fetchall()
    return [_historico_dict(r) for r in rows]


def relatorio_atividade_fiscal(base_id, limite=40):
    """Últimas ações do fiscal na filial (para relatórios gerenciais)."""
    init_db()
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT * FROM historico_ativos
            WHERE base_id = ?
            ORDER BY criado_em DESC, id DESC
            LIMIT ?
            """,
            (int(base_id), int(limite)),
        ).fetchall()
    return [_historico_dict(r) for r in rows]


def atualizar_historico(historico_id, dados):
    """Somente admin: corrige um registro do histórico."""
    atual = obter_historico(historico_id)
    if not atual:
        raise ValueError("Registro de histórico não encontrado.")

    codigo = (dados.get("codigo") or atual["codigo"] or "").strip() or "—"
    nome = (dados.get("nome") or atual["nome"] or "").strip()
    if not nome:
        raise ValueError("Informe o nome do ativo no histórico.")

    acao = (dados.get("acao") or atual["acao"] or "verificacao").strip()
    if acao not in ACAO_LABEL:
        raise ValueError("Ação de histórico inválida.")

    detalhes = (dados.get("detalhes") or "").strip()
    ordem_servico = (dados.get("ordem_servico") or "").strip()
    local_raw = (dados.get("local") or "").strip().lower()
    local = local_raw if local_raw in LOCAIS_VALIDOS else (atual.get("local") or "base")

    em_manutencao = dados.get("em_manutencao")
    if em_manutencao is None:
        em_manutencao = atual.get("em_manutencao")
    else:
        em_manutencao = bool(em_manutencao)

    with get_conn() as conn:
        conn.execute(
            """
            UPDATE historico_ativos
            SET codigo = ?, nome = ?, acao = ?, detalhes = ?,
                em_manutencao = ?, local = ?, ordem_servico = ?
            WHERE id = ?
            """,
            (
                codigo,
                nome,
                acao,
                detalhes,
                int(em_manutencao) if em_manutencao is not None else None,
                local,
                ordem_servico,
                int(historico_id),
            ),
        )

    return obter_historico(historico_id)


def excluir_historico(historico_id):
    """Somente admin: remove um registro do histórico."""
    atual = obter_historico(historico_id)
    if not atual:
        raise ValueError("Registro de histórico não encontrado.")

    with get_conn() as conn:
        conn.execute("DELETE FROM historico_ativos WHERE id = ?", (int(historico_id),))

    return atual


# --- Admin ---


def autenticar_admin(usuario, senha):
    init_db()
    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT * FROM admins
            WHERE usuario = ? AND ativo = 1
            """,
            (usuario.strip().lower(),),
        ).fetchone()

    if row and check_password_hash(row["senha_hash"], senha):
        d = dict(row)
        d["id"] = str(d["id"])
        d["ativo"] = bool(d["ativo"])
        return d
    return None


def obter_admin(admin_id):
    init_db()
    with get_conn() as conn:
        row = conn.execute("SELECT * FROM admins WHERE id = ?", (int(admin_id),)).fetchone()
    if not row:
        return None
    d = dict(row)
    d["id"] = str(d["id"])
    d["ativo"] = bool(d["ativo"])
    return d


def listar_todos_usuarios(base_id=None):
    init_db()
    sql = """
        SELECT u.*, b.codigo AS base_codigo, b.nome AS base_nome
        FROM usuarios u
        JOIN bases b ON b.id = u.base_id
    """
    params = []
    if base_id:
        sql += " WHERE u.base_id = ?"
        params.append(int(base_id))
    sql += " ORDER BY b.codigo, u.nivel, u.usuario"

    with get_conn() as conn:
        rows = conn.execute(sql, params).fetchall()

    result = []
    for row in rows:
        d = _row_to_dict(row)
        d["base_codigo"] = row["base_codigo"]
        d["base_nome"] = row["base_nome"]
        result.append(d)
    return result


def obter_usuario_por_id(user_id):
    init_db()
    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT u.*, b.codigo AS base_codigo, b.nome AS base_nome
            FROM usuarios u
            JOIN bases b ON b.id = u.base_id
            WHERE u.id = ?
            """,
            (int(user_id),),
        ).fetchone()
    if not row:
        return None
    d = _row_to_dict(row)
    d["base_codigo"] = row["base_codigo"]
    d["base_nome"] = row["base_nome"]
    return d


def admin_resetar_senha(user_id, nova_senha):
    if len(nova_senha) < 4:
        raise ValueError("A senha deve ter pelo menos 4 caracteres.")
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE usuarios
            SET senha_hash = ?, senha_alterada = 0
            WHERE id = ?
            """,
            (generate_password_hash(nova_senha), int(user_id)),
        )


def admin_toggle_usuario(user_id, ativo):
    with get_conn() as conn:
        conn.execute(
            "UPDATE usuarios SET ativo = ? WHERE id = ?",
            (int(ativo), int(user_id)),
        )


def admin_atualizar_usuario(user_id, nome, nivel):
    if nivel not in ("gerente", "fiscal"):
        raise ValueError("Nível inválido. Use gerente ou fiscal.")
    if not nome.strip():
        raise ValueError("Nome é obrigatório.")
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE usuarios SET nome = ?, nivel = ?
            WHERE id = ?
            """,
            (nome.strip(), nivel, int(user_id)),
        )


def admin_criar_usuario(base_id, usuario, nome, nivel, senha):
    usuario = usuario.strip().lower()
    if not usuario or not nome.strip():
        raise ValueError("Usuário e nome são obrigatórios.")
    if nivel not in ("gerente", "fiscal"):
        raise ValueError("Nível inválido. Use gerente ou fiscal.")
    if len(senha) < 4:
        raise ValueError("A senha deve ter pelo menos 4 caracteres.")
    try:
        with get_conn() as conn:
            conn.execute(
                """
                INSERT INTO usuarios
                    (base_id, usuario, senha_hash, nome, nivel, senha_alterada, ativo)
                VALUES (?, ?, ?, ?, ?, 0, 1)
                """,
                (
                    int(base_id),
                    usuario,
                    generate_password_hash(senha),
                    nome.strip(),
                    nivel,
                ),
            )
    except sqlite3.IntegrityError as exc:
        raise ValueError("Já existe este usuário nesta filial.") from exc


def admin_alterar_senha_admin(admin_id, senha_atual, nova_senha):
    admin = obter_admin(admin_id)
    if not admin or not check_password_hash(admin["senha_hash"], senha_atual):
        raise ValueError("Senha atual incorreta.")
    if len(nova_senha) < 4:
        raise ValueError("A nova senha deve ter pelo menos 4 caracteres.")
    with get_conn() as conn:
        conn.execute(
            """
            UPDATE admins SET senha_hash = ? WHERE id = ?
            """,
            (generate_password_hash(nova_senha), int(admin_id)),
        )

from equipamentos import (  # noqa: E402
    abrir_manutencao,
    adicionar_anotacao_manutencao,
    atualizar_dados_os_aberta,
    salvar_acompanhamento_os_aberta,
    atualizar_fase_manutencao,
    encerrar_manutencao,
    equipamento_detalhe,
    export_csv_manutencoes,
    listar_ativos_com_stats,
    listar_manutencoes_ativo,
    listar_manutencoes_base,
    obter_manutencao_aberta,
    relatorio_manutencoes,
    stats_ativo,
)
